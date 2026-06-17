import os
from openpyxl import load_workbook


def read_excel_to_2d_array(file_path: str) -> list:
    """
    读取Excel文件内容，返回二维数组
    每一行是一个数组，所有行组成一个顶级数组
    
    参数:
        file_path: Excel文件路径
        
    返回:
        二维数组，如 [[cell1, cell2], [cell3, cell4], ...]
    """
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return []
    
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        result = []
        for row in ws.iter_rows(values_only=True):
            row_data = []
            for cell in row:
                row_data.append(cell)
            result.append(row_data)
        
        wb.close()
        return result
    except Exception as e:
        print(f"读取Excel失败: {e}")
        return []

def copy_and_write_sheets_from_template(data: list, template_path: str = None, output_path: str = None) -> bool:
    """
    以记录表模板.xlsx的第一个sheet为模板，根据二维数组的行数复制sheet
    
    参数:
        data: 二维数组，每行数据对应一个sheet
        template_path: 模板文件路径，默认为项目根目录下的 记录表模板.xlsx
        output_path: 输出文件路径，默认为项目根目录下的 记录表输出.xlsx
        
    返回:
        成功返回True，失败返回False
    """
    if not data:
        print("数据为空")
        return False
    
    project_dir = os.path.dirname(os.path.abspath(__file__))
    
    if template_path is None:
        template_path = os.path.join(project_dir, '记录表.xlsx')
    
    if output_path is None:
        output_path = os.path.join(project_dir, '记录表输出.xlsx')
    
    if not os.path.exists(template_path):
        print(f"模板文件不存在: {template_path}")
        return False
    
    try:
        wb = load_workbook(template_path)
        
        if not wb.sheetnames:
            print("模板文件没有sheet")
            wb.close()
            return False
        
        template_sheet = wb[wb.sheetnames[0]]
        qd="C35"
        for i, row in enumerate(data):
            sheet_name = f"sheet{i+1}"
            
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            
            wb.copy_worksheet(template_sheet)
            new_sheet = wb.worksheets[-1]
            new_sheet.title = sheet_name

            tdz=data[i][5] # 推定值
            pjz_10=data[i][2] # 10点平均值
            zxz=data[i][4] # 最小值
            bzc=data[i][3] # 标准差
            
            if "阁楼层梁" in data[i][0]: qd = "C30"

            
            new_sheet.cell(row=2, column=25, value=f"第{i+1}页/共30页")
            new_sheet.cell(row=4, column=21, value=data[i][0]) # 构件名称
            # 推定值:[ 31.1MPa ] 平均值:[ 40.5MPa ] 最小值:[ 32.0MPa ] 标准差:[ 5.69MPa ]
            new_sheet.cell(row=22, column=3, value=f"推定值:[ {tdz}MPa ] 平均值:[ {pjz_10}MPa ] 最小值:[ {zxz}MPa ] 标准差:[ {bzc}MPa ]")
            new_sheet.cell(row=19, column=8, value=qd) # 强度
            


            for x in range(7,17):
                new_sheet.cell(row=x, column=19, value=data[i][(x-7)*18+21]) # 16点平均值,22,185
                new_sheet.cell(row=x, column=26, value=data[i][(x-7)*18+22]) # 转换值,23
                print(len(data[i]))
                for y in range(3, 19):
                    new_sheet.cell(row=x, column=y, value=data[i][(x-7)*18+5+y-3]) # 实测值








        wb.save(output_path)
        wb.close()
        return True
    except Exception as e:
        print(f"复制sheet失败: {e}")
        return False


"""
# 方式1：使用单元格坐标（行号，列号）
ws.cell(row=1, column=1, value="内容")   # 写入A1单元格
ws.cell(row=5, column=3, value=123.45)   # 写入C5单元格
"""


if __name__ == '__main__':
    project_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(project_dir, '写到记录表的数据.xlsx')
    
    data = read_excel_to_2d_array(excel_path)
    
    print(f"总行数: {len(data)}")
    

    excel_path2 = os.path.join(project_dir, '记录表.xlsx')
    success=copy_and_write_sheets_from_template(data, excel_path2)
    print(success)
    

